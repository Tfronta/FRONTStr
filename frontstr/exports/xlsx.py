"""Multi-sheet XLSX export.

The CSV exports are the machine-readable ones; this is the human-readable one.
It exists because forensic review happens in Excel, and handing an analyst five
separate CSVs and asking them to line up the marker columns is how transcription
errors get made.

Five sheets, in the order a review actually proceeds:

1. **Profile** — the genotype table. One row per marker, allele numbers and
   per-allele coverage. This is what gets compared against a reference profile.
2. **Sequences** — one row per called allele: ISFG string, iso-allele, the full
   consensus. The sequence-level evidence CE cannot provide.
3. **Evidence** — every cluster, called or not, including the ones classified
   as stutter or suppressed as haplotype phantoms. This is where you look when
   you want to know *why* a call came out the way it did.
4. **QC** — one row per flag, with severity and the full message.
5. **Audit** — the run configuration and the flag census, as key/value.

Markers carrying a warning or error are tinted on the Profile sheet, and the QC
sheet is ordered by severity, so the review starts where the problems are
instead of at D3S1358.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from frontstr.interp.isfg import motif_repeat_summary

_HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_WARN_FILL = PatternFill("solid", fgColor="FFF3CD")
_ERROR_FILL = PatternFill("solid", fgColor="F8D7DA")
_MONO = Font(name="Menlo", size=10)

#: Long sequence columns are capped at this width; the value is intact in the
#: cell, only the display is narrowed. Without this a 250 bp consensus makes the
#: sheet unusable.
_MAX_COL_WIDTH = 60

_SEVERITY_ORDER = {"error": 0, "warn": 1, "info": 2}


def _write_sheet(
    ws: Worksheet,
    headers: list[str],
    rows: list[list[Any]],
    *,
    mono_columns: set[str] | None = None,
    row_fills: dict[int, PatternFill] | None = None,
) -> None:
    """Write a header + rows with a frozen header, autofilter and sane widths."""
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")

    for row in rows:
        ws.append(row)

    ws.freeze_panes = "A2"
    # Filter the header even with no data rows: an empty QC sheet that cannot be
    # filtered looks broken next to four that can.
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    mono = mono_columns or set()
    for idx, name in enumerate(headers, start=1):
        letter = get_column_letter(idx)
        longest = max(
            [len(str(name))] + [len(str(r[idx - 1])) for r in rows if r[idx - 1] is not None]
        )
        ws.column_dimensions[letter].width = min(longest + 2, _MAX_COL_WIDTH)
        if name in mono:
            for cell in ws[letter][1:]:
                cell.font = _MONO

    for row_idx, fill in (row_fills or {}).items():
        for cell in ws[row_idx]:
            cell.fill = fill


def _severity_of(marker: dict[str, Any]) -> str | None:
    """Worst severity carried by a marker, or ``None`` when clean."""
    severities = [f["severity"] for f in marker.get("flags", [])]
    for level in ("error", "warn"):
        if level in severities:
            return level
    return None


def _profile_sheet(ws: Worksheet, payload: dict[str, Any]) -> None:
    headers = [
        "marker",
        "genotype",
        "call_rule",
        "coverage",
        "allele1",
        "allele1_reads",
        "allele2",
        "allele2_reads",
        "allele3",
        "allele3_reads",
        "iso_allele",
        "flags",
    ]
    rows: list[list[Any]] = []
    fills: dict[int, PatternFill] = {}

    for marker in payload.get("results", []):
        called = marker.get("alleles_called", [])
        labels = [a.get("number_label") or "?" for a in called]
        row: list[Any] = [
            marker["marker_name"],
            " / ".join(labels) if labels else "—",
            marker["call_rule"],
            marker["total_reads"],
        ]
        for i in range(3):
            if i < len(called):
                row += [labels[i], called[i]["n_reads_total"]]
            else:
                row += [None, None]
        row.append(
            ", ".join(
                (a.get("iso") or {}).get("suffix") or ""
                for a in called
                if (a.get("iso") or {}).get("is_isoallele")
            )
            or None
        )
        row.append(", ".join(f["code"] for f in marker.get("flags", [])) or None)
        rows.append(row)

        severity = _severity_of(marker)
        if severity:
            fills[len(rows) + 1] = _ERROR_FILL if severity == "error" else _WARN_FILL

    _write_sheet(ws, headers, rows, row_fills=fills)


def _sequences_sheet(ws: Worksheet, payload: dict[str, Any]) -> None:
    headers = [
        "marker",
        "allele_index",
        "allele_number",
        "number_method",
        "isfg",
        "repeat_summary",
        "iso_suffix",
        "iso_match",
        "length_bp",
        "reads",
        "reads_hp1",
        "reads_hp2",
        "consensus_method",
        "consensus",
    ]
    rows: list[list[Any]] = []
    for marker in payload.get("results", []):
        motif = marker["system"]["motif"]
        strand = marker["system"].get("strand", "+")
        for i, a in enumerate(marker.get("alleles_called", []), start=1):
            iso = a.get("iso") or {}
            rows.append(
                [
                    marker["marker_name"],
                    i,
                    a.get("number_label"),
                    a.get("number_method"),
                    a["isfg"],
                    motif_repeat_summary(a["consensus"], motif, strand=strand),
                    iso.get("suffix"),
                    iso.get("match_type") if iso.get("suffix") else None,
                    a["length_bp"],
                    a["n_reads_total"],
                    a["n_reads_hp1"],
                    a["n_reads_hp2"],
                    a.get("consensus_method"),
                    a["consensus"],
                ]
            )
    _write_sheet(ws, headers, rows, mono_columns={"isfg", "consensus"})


def _evidence_sheet(ws: Worksheet, payload: dict[str, Any]) -> None:
    headers = [
        "marker",
        "cluster",
        "status",
        "called",
        "length_bp",
        "reads",
        "fraction",
        "reads_hp1",
        "reads_hp2",
        "reads_untagged",
        "forward",
        "reverse",
        "mean_qual",
        "expected_stutter",
        "reads_absorbed",
        "isfg",
    ]
    rows: list[list[Any]] = []
    for marker in payload.get("results", []):
        called_ids = {a["cluster_index"] for a in marker.get("alleles_called", [])}
        for a in marker.get("alleles", []):
            rows.append(
                [
                    marker["marker_name"],
                    a["cluster_index"],
                    a["status"],
                    # None, not "": Excel reads a blank string back as empty
                    # anyway, and a genuinely empty cell filters correctly.
                    "yes" if a["cluster_index"] in called_ids else None,
                    a["length_bp"],
                    a["n_reads_total"],
                    a["fraction"],
                    a["n_reads_hp1"],
                    a["n_reads_hp2"],
                    a["n_reads_hp_none"],
                    a["n_forward"],
                    a["n_reverse"],
                    a["mean_qual"],
                    a["expected_stutter"],
                    a.get("n_reads_absorbed", 0),
                    a["isfg"],
                ]
            )
    _write_sheet(ws, headers, rows, mono_columns={"isfg"})


def _qc_sheet(ws: Worksheet, payload: dict[str, Any]) -> None:
    headers = ["severity", "marker", "code", "message"]
    rows: list[list[Any]] = []
    for marker in payload.get("results", []):
        for f in marker.get("flags", []):
            rows.append([f["severity"], marker["marker_name"], f["code"], f["message"]])
        for a in marker.get("alleles", []):
            for f in a.get("flags", []):
                rows.append(
                    [
                        f["severity"],
                        f"{marker['marker_name']} (cluster {a['cluster_index']})",
                        f["code"],
                        f["message"],
                    ]
                )
    # Worst first: a reviewer should not have to scroll to find the errors.
    rows.sort(key=lambda r: (_SEVERITY_ORDER.get(str(r[0]), 9), str(r[1])))

    fills = {
        i + 2: (_ERROR_FILL if r[0] == "error" else _WARN_FILL)
        for i, r in enumerate(rows)
        if r[0] in ("error", "warn")
    }
    _write_sheet(ws, headers, rows, row_fills=fills)


def _audit_sheet(ws: Worksheet, payload: dict[str, Any]) -> None:
    meta = payload.get("meta", {})
    audit = payload.get("audit", {})
    qc = audit.get("qc_thresholds", {})

    rows: list[list[Any]] = [
        ["Sample", meta.get("sample_name")],
        ["Operator", meta.get("operator")],
        ["Run ID", meta.get("run_id")],
        ["Started at", meta.get("started_at")],
        ["", ""],
        ["FRONTStr version", audit.get("tool_version")],
        ["POA backend", audit.get("poa_backend")],
        ["Stutter model", audit.get("stutter_model_version")],
        ["Stutter protocol", audit.get("stutter_model_protocol")],
        ["", ""],
        ["Panel", f"{meta.get('panel_name', '')} {meta.get('panel_version', '')}".strip()],
        ["Reference build", meta.get("reference_build")],
        ["Analytical threshold", audit.get("analytical_thresh")],
        ["Calling threshold", audit.get("calling_thresh")],
        ["Low-coverage floor (reads)", qc.get("low_coverage_reads")],
        ["Strand-bias p", qc.get("strand_bias_p")],
        ["Strand-bias min reads", qc.get("strand_bias_min_reads")],
        ["", ""],
    ]
    for item in audit.get("inputs", []):
        rows.append([f"Input ({item.get('role')})", item.get("path")])
        rows.append([f"  sha256 ({item.get('role')})", item.get("sha256")])
    rows.append(["", ""])

    counts = audit.get("flag_counts", {})
    for code in audit.get("flags_checked", []):
        rows.append([f"Flag: {code}", counts.get(code, 0)])
    rows.append(["", ""])
    rows.append(["Markers needing review", ", ".join(audit.get("markers_needing_review", []))])
    rows.append(["Audit record sha256", audit.get("integrity_sha256")])

    _write_sheet(ws, ["Field", "Value"], rows)
    # A key/value sheet has nothing meaningful to filter.
    ws.auto_filter.ref = None


def write_run_xlsx(payload: dict[str, Any], out_path: Path) -> Path:
    """Write the multi-sheet review workbook.

    Args:
        payload: Output of :func:`frontstr.report.payload.serialize_run`.
        out_path: Destination ``.xlsx``.

    Returns:
        ``out_path``.
    """
    wb = Workbook()
    # A new Workbook always has one default sheet; drop it so the five below
    # are the whole workbook and appear in review order.
    default = wb.active
    if default is not None:
        wb.remove(default)

    for title, builder in (
        ("Profile", _profile_sheet),
        ("Sequences", _sequences_sheet),
        ("Evidence", _evidence_sheet),
        ("QC", _qc_sheet),
        ("Audit", _audit_sheet),
    ):
        builder(wb.create_sheet(title), payload)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def write_cohort_xlsx(cohort: dict[str, Any], out_path: Path) -> Path:
    """Write the cohort workbook: one row per sample per marker, plus overviews.

    Three sheets, in the order a reviewer wants them:

    ``Genotypes``
        The long form. One row per (marker, sample) with both alleles, their
        read counts, the coverage of the call, allele balance, the QC codes and
        both ISFG strings. Long rather than wide because that is the shape a
        spreadsheet can filter, sort and pivot; a 25-marker-wide sheet cannot
        be filtered by "show me the flagged calls".
    ``By marker`` / ``By sample``
        The two margins, so the question "which locus is failing" and "which
        sample is failing" are each one glance rather than a pivot table.

    Args:
        cohort: Output of :func:`frontstr.report.cohort.build_cohort_payload`.
        out_path: Destination ``.xlsx``.
    """
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    headers = [
        "Marker",
        "Sample",
        "Call",
        "Allele 1",
        "Reads 1",
        "Allele 2",
        "Reads 2",
        "Cov",
        "Discarded",
        "AB",
        "QC",
        "ISFG allele 1",
        "ISFG allele 2",
    ]
    rows: list[list[Any]] = []
    fills: dict[int, PatternFill] = {}
    for block in cohort["blocks"]:
        for row in block.rows:
            rows.append(
                [
                    block.marker,
                    row.get("sample", ""),
                    row.get("call_rule", ""),
                    row.get("allele1_ce_label"),
                    row.get("allele1_cov"),
                    row.get("allele2_ce_label"),
                    row.get("allele2_cov"),
                    row.get("called_reads"),
                    row.get("discarded_reads"),
                    row.get("allele_balance"),
                    ", ".join(f.get("code", "") for f in row.get("flags", [])),
                    row.get("allele1_isfg"),
                    row.get("allele2_isfg"),
                ]
            )
            severity = row.get("worst_severity")
            if severity in ("warn", "error"):
                # +1 for the header row, and the row was just appended.
                fills[len(rows) + 1] = _ERROR_FILL if severity == "error" else _WARN_FILL

    _write_sheet(
        wb.create_sheet("Genotypes"),
        headers,
        rows,
        mono_columns={"ISFG allele 1", "ISFG allele 2"},
        row_fills=fills,
    )

    _write_sheet(
        wb.create_sheet("By marker"),
        ["Marker", "Samples", "Called", "Call rate", "Flagged calls", "Flags raised"],
        [
            [
                b.marker,
                b.n_samples,
                b.n_called,
                round(b.call_rate, 4),
                b.n_flagged,
                ", ".join(f"{code} ×{n}" for _short, code, n in b.flag_counts),
            ]
            for b in cohort["blocks"]
        ],
    )

    _write_sheet(
        wb.create_sheet("By sample"),
        ["Sample", "Markers", "Called", "Call rate", "Flagged calls"],
        [
            [s.sample_id, s.n_markers, s.n_called, round(s.call_rate, 4), s.n_flagged]
            for s in cohort["samples"]
        ],
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
