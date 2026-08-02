"""Read the external truth genotypes out of ``1000GEN-ONT-Merged-Compar.xlsx``.

The workbook is not part of this repository and its layout is bespoke, so this
module is deliberately dev tooling under ``benchmark/`` rather than part of the
``frontstr`` package. It converts the spreadsheet into a tidy table that the
comparator can consume, so nothing downstream has to know about merged header
cells or Spanish column names.

Only the public samples in that workbook have any of this. A user calling their
own ONT samples has no truth table and needs none — see :mod:`benchmark`.

Sheet ``Concordancia-3-tecnologias`` layout:

===  ==========================================================
row  contents
===  ==========================================================
0    marker name, written once per block into a merged cell
1    technology — ``ILLUMINA`` | ``longTR`` | ``STRspy``
2    field — ``A1`` | ``A2`` | ``COBERTURA`` | ``COB A1`` | ...
3+   one row per sample; column 0 is the sample ID
===  ==========================================================

Marker blocks are *not* a fixed width — STRspy carries two coverage columns for
some markers and one for others, and only some blocks have a trailing
``CONCORDANCIA`` column. Parsing therefore keys on the (marker, technology,
field) triple read out of the three header rows, never on a column offset.
"""

from __future__ import annotations

from collections.abc import Sequence
from dataclasses import dataclass
from pathlib import Path

import openpyxl

TRUTH_SHEET = "Concordancia-3-tecnologias"

#: The three genotype sources the sheet carries, in adjudication order:
#: Illumina is the only non-ONT technology and so the only external anchor.
TECHNOLOGIES = ("ILLUMINA", "longTR", "STRspy")

#: Markers that appear in the workbook but not in FRONTStr's panel.
NOT_IN_PANEL = frozenset({"PentaD", "PentaE", "SE30"})

#: Markers with no Illumina column anywhere in the workbook. Their only
#: reference is longTR + STRspy, i.e. two ONT callers on the same data, so a
#: FRONTStr agreement there is caller-vs-caller and not external validation.
#: See the ``reference-frontstr-truth-workbook`` note.
NO_EXTERNAL_TRUTH = frozenset({"D21S11"})

#: Motif length per marker, needed to read a cell written as bases over motif
#: length instead of in ISFG notation. Held here rather than read from the
#: panel because the workbook is a fixed artifact and the benchmark must be
#: able to parse it without one; ``test_motif_lengths_match_the_shipped_panel``
#: is what stops the two from drifting apart.
MOTIF_LENGTHS = {
    "CSF1PO": 4,
    "D10S1248": 4,
    "D12S391": 4,
    "D13S317": 4,
    "D16S539": 4,
    "D18S51": 4,
    "D19S433": 4,
    "D1S1656": 4,
    "D21S11": 4,
    "D22S1045": 3,
    "D2S1338": 4,
    "D2S441": 4,
    "D3S1358": 4,
    "D5S818": 4,
    "D7S820": 4,
    "D8S1179": 4,
    "FGA": 4,
    "TH01": 4,
    "TPOX": 4,
    "vWA": 4,
}

#: A genotype: one or two allele designations as canonical strings.
Genotype = tuple[str, ...]


@dataclass(frozen=True, slots=True)
class TruthCall:
    """One (sample, marker, technology) genotype from the workbook."""

    sample_id: str
    marker: str
    technology: str
    genotype: Genotype
    coverage: float | None = None

    @property
    def is_missing(self) -> bool:
        """True when this technology reported nothing for the locus."""
        return not self.genotype


def canonical_sample_id(raw: str) -> str:
    """Normalise a sample ID to the workbook's spelling.

    The aligned BAM filenames use a ``GM`` prefix where the workbook (and
    1000 Genomes) uses ``NA`` — ``GM19038`` is ``NA19038``. ``HG`` IDs are
    spelled the same on both sides.
    """
    sid = raw.strip().upper()
    if sid.startswith("GM"):
        return "NA" + sid[2:]
    return sid


#: How far a stored suffix may sit from a notation's exact value and still be
#: recognised as it. A third written to four decimals is off by 3e-5, and a
#: typed cell can be shorter still. The nearest two candidates anywhere in the
#: panel are ``.2`` and ``.25``, so anything well under 0.025 separates every
#: pair; at 0.01 a suffix like ``.45``, which is neither notation, still falls
#: through to being reported rather than guessed at.
_SUFFIX_TOL = 0.01


def isfg_from_repeat_units(value: float, motif_length: int) -> str | None:
    """ISFG designation for a cell written in repeat units, or ``None``.

    HipSTR and longTR report an allele in **bases**, and the workbook divides
    that by the motif length, so three extra bases on a tetranucleotide are
    stored as ``.75`` and not as ISFG's ``.3``. The two notations are
    distinguishable: an ISFG suffix counts extra bases and is written as a
    tenth, so it is one of ``.1 .2 .3`` on a tetranucleotide, while the
    quotient can only be one of ``.25 .5 .75``. For every marker in this panel,
    which is tri- and tetranucleotide throughout, those two sets are disjoint.

    They stop being disjoint at a pentanucleotide, where ``.2`` and ``.4`` are
    valid in both notations and mean different alleles. Nothing is converted
    there; guessing would be worse than reporting the cell as it stands.

    Returns:
        The ISFG designation when ``value`` is unambiguously in repeat units,
        ``None`` when it is already ISFG, or ambiguous, or neither.
    """
    if motif_length >= 5:
        return None
    fraction = value % 1
    if fraction < _SUFFIX_TOL or fraction > 1 - _SUFFIX_TOL:
        return None
    if any(abs(fraction - extra / 10) < _SUFFIX_TOL for extra in range(1, motif_length)):
        return None
    for extra in range(1, motif_length):
        if abs(fraction - extra / motif_length) < _SUFFIX_TOL:
            return f"{int(value)}.{extra}"
    return None


def canonical_allele(
    raw: object,
    *,
    motif_length: int | None = None,
    converted: list[tuple[float, str]] | None = None,
) -> str | None:
    """Render one allele cell as a canonical CE string, or ``None`` if absent.

    Accepts the several spellings the workbook and FRONTStr disagree on:
    ``13`` (int), ``13.0`` (float), ``"13"``, ``"9.3"``. Empty cells and the
    workbook's ``NA`` / ``ND`` placeholders become ``None``.

    Whole numbers lose their ``.0`` so that FRONTStr's ``"12.0"`` and the
    workbook's ``12`` compare equal.

    Args:
        raw: The cell.
        motif_length: Set for workbook cells, which may be written in repeat
            units and need :func:`isfg_from_repeat_units`. Left unset for
            FRONTStr's own output, which is ISFG already and must not be
            reinterpreted.
        converted: Optional list appended with ``(stored value, ISFG)`` for
            every cell that was in repeat units, so the run can report what it
            re-read rather than silently changing a genotype.
    """
    if raw is None:
        return None
    text = str(raw).strip()
    if not text or text.upper() in {"NA", "ND", "-", "?", "N/A"}:
        return None
    try:
        value = float(text)
    except ValueError:
        # Something like "16/18" or a stray note — keep it verbatim so the
        # mismatch is visible rather than silently dropped.
        return text
    if motif_length is not None:
        isfg = isfg_from_repeat_units(value, motif_length)
        if isfg is not None:
            if converted is not None:
                converted.append((value, isfg))
            return isfg
    # Rounding here is what made a stored 9.25 read as "9.2", an allele the
    # workbook never asserted and one FRONTStr can legitimately call. Anything
    # that reaches this line is already a tenth.
    return f"{value:.1f}".removesuffix(".0")


def canonical_genotype(
    alleles: list[object],
    *,
    motif_length: int | None = None,
    converted: list[tuple[float, str]] | None = None,
) -> Genotype:
    """Build a comparable genotype from raw allele cells.

    A single reported allele is a homozygote, so it is expanded to a pair: a
    FRONTStr homozygous call renders as one CE value while the workbook writes
    the same allele into both A1 and A2, and those two must compare equal.
    Alleles are sorted, because A1/A2 order is not meaningful and the workbook
    is not consistent about it (``13, 12`` appears alongside ``10, 12``).
    """
    called = [
        a
        for a in (
            canonical_allele(x, motif_length=motif_length, converted=converted) for x in alleles
        )
        if a is not None
    ]
    if not called:
        return ()
    if len(called) == 1:
        called = called * 2
    return tuple(sorted(called, key=_allele_sort_key))


def _allele_sort_key(allele: str) -> tuple[float, str]:
    """Sort alleles numerically where possible, lexically otherwise."""
    try:
        return (float(allele), "")
    except ValueError:
        return (float("inf"), allele)


def _header_map(rows: Sequence[tuple[object, ...]]) -> dict[tuple[str, str, str], int]:
    """Map (marker, technology, field) to its column index.

    Rows 0 and 1 are *both* merged: the marker name is written once per block
    and the technology name once per sub-block, leaving every following cell
    empty. Both are therefore forward-filled. Reading only the non-empty cells
    would silently keep A1 and drop A2, which does not fail — it turns every
    genotype into a homozygote.

    The technology fill resets at each marker boundary so that a trailing
    column (the ungrouped ``CONCORDANCIA``) cannot leak the previous block's
    technology into the next marker.
    """
    marker_row, tech_row, field_row = rows[0], rows[1], rows[2]
    width = max(len(marker_row), len(tech_row), len(field_row))

    columns: dict[tuple[str, str, str], int] = {}
    current_marker: str | None = None
    current_tech: str | None = None
    for col in range(width):
        if col < len(marker_row) and marker_row[col]:
            current_marker = str(marker_row[col]).strip()
            current_tech = None
        if col < len(tech_row) and tech_row[col]:
            current_tech = str(tech_row[col]).strip()
        if current_marker is None or current_tech is None:
            continue
        field = field_row[col] if col < len(field_row) else None
        if field:
            columns[(current_marker, current_tech, str(field).strip())] = col
    return columns


def load_truth(
    workbook: Path,
    *,
    sheet: str = TRUTH_SHEET,
    include_not_in_panel: bool = False,
    converted: list[tuple[str, str, str, float, str]] | None = None,
) -> list[TruthCall]:
    """Read every (sample, marker, technology) genotype out of the workbook.

    Args:
        workbook: Path to ``1000GEN-ONT-Merged-Compar.xlsx``.
        sheet: Sheet to read. The default carries all three technologies side
            by side, which is what makes a disagreement adjudicable.
        include_not_in_panel: Keep PentaD/PentaE, which FRONTStr does not call.
        converted: Optional list appended with ``(sample, marker, technology,
            stored value, ISFG)`` for every cell that was written in repeat
            units. Re-reading a genotype is not something a validation run gets
            to do quietly, so the caller is expected to report these.

    Returns:
        One :class:`TruthCall` per sample × marker × technology, including the
        ones where that technology reported nothing (``is_missing``), so that a
        caller can tell "no truth" apart from "not looked at".
    """
    if not workbook.exists():
        raise FileNotFoundError(f"Truth workbook not found: {workbook}")

    wb = openpyxl.load_workbook(workbook, data_only=True, read_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise KeyError(f"Sheet {sheet!r} not in {workbook.name}; has {wb.sheetnames}")
        ws = wb[sheet]
        rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    if len(rows) < 4:
        raise ValueError(f"Sheet {sheet!r} has no data rows")

    columns = _header_map(rows[:3])
    markers = sorted({marker for marker, _, _ in columns})
    if not include_not_in_panel:
        markers = [m for m in markers if m not in NOT_IN_PANEL]

    calls: list[TruthCall] = []
    for row in rows[3:]:
        if not row or not row[0]:
            continue
        sample_id = canonical_sample_id(str(row[0]))
        if not sample_id[:2].isalpha():
            continue

        for marker in markers:
            for tech in TECHNOLOGIES:
                a1 = _cell(row, columns.get((marker, tech, "A1")))
                a2 = _cell(row, columns.get((marker, tech, "A2")))
                cells: list[tuple[float, str]] = []
                genotype = canonical_genotype(
                    [a1, a2],
                    motif_length=MOTIF_LENGTHS.get(marker),
                    converted=cells,
                )
                if converted is not None:
                    converted.extend(
                        (sample_id, marker, tech, stored, isfg) for stored, isfg in cells
                    )
                calls.append(
                    TruthCall(
                        sample_id=sample_id,
                        marker=marker,
                        technology=tech,
                        genotype=genotype,
                        coverage=_coverage(row, columns, marker, tech),
                    )
                )
    return calls


def _cell(row: tuple[object, ...], col: int | None) -> object:
    """Value at ``col``, or ``None`` when the column is absent or short."""
    if col is None or col >= len(row):
        return None
    return row[col]


def _coverage(
    row: tuple[object, ...],
    columns: dict[tuple[str, str, str], int],
    marker: str,
    tech: str,
) -> float | None:
    """Read the coverage cell, whichever of the two spellings this block uses.

    STRspy blocks carry per-allele coverage (``COB A1`` / ``COB A2``); the
    others carry one ``COBERTURA``. Per-allele coverage is summed so that the
    single number means the same thing across technologies: reads at the locus.
    """
    single = _cell(row, columns.get((marker, tech, "COBERTURA")))
    if single is not None:
        return _as_float(single)

    parts = [
        _as_float(_cell(row, columns.get((marker, tech, field)))) for field in ("COB A1", "COB A2")
    ]
    present = [p for p in parts if p is not None]
    return sum(present) if present else None


def _as_float(raw: object) -> float | None:
    """Parse a numeric cell, tolerating ``NA`` and blanks."""
    if raw is None:
        return None
    try:
        return float(str(raw).strip())
    except ValueError:
        return None


def index_by_sample_marker(
    calls: list[TruthCall],
) -> dict[tuple[str, str], dict[str, TruthCall]]:
    """Group truth calls as ``{(sample, marker): {technology: call}}``."""
    index: dict[tuple[str, str], dict[str, TruthCall]] = {}
    for call in calls:
        index.setdefault((call.sample_id, call.marker), {})[call.technology] = call
    return index
