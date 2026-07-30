"""Tests for the cohort view: many samples, one block per marker.

The single-sample report answers "is this profile right". A hundred samples ask
different questions, and this view exists to answer them, so the tests are about
whether it points somewhere rather than about whether it renders.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import lxml.html
import pytest

from frontstr.errors import FrontstrError
from frontstr.report.cohort import build_cohort_payload, build_cohort_report


def _row(
    sample: str,
    marker: str,
    *,
    a1: str | None = "14",
    a2: str | None = "16",
    cov: int = 33,
    flags: list[dict[str, str]] | None = None,
) -> dict[str, Any]:
    return {
        "sample": sample,
        "marker": marker,
        "allele1_ce_label": a1,
        "allele1_cov": 17 if a1 else None,
        "allele1_isfg": f"CE{a1}_TATC[10]" if a1 else None,
        "allele2_ce_label": a2,
        "allele2_cov": 16 if a2 else None,
        "allele2_isfg": f"CE{a2}_TATC[12]" if a2 else None,
        "called_reads": cov,
        "discarded_reads": 5,
        "total_reads": cov + 5,
        "allele_balance": 0.52,
        "flags": flags or [],
        "worst_severity": flags[0]["severity"] if flags else None,
    }


def _payload(sample: str, rows: list[dict[str, Any]]) -> dict[str, Any]:
    return {"meta": {"sample_name": sample}, "profile_rows": rows}


def _lc(severity: str = "warn") -> list[dict[str, str]]:
    return [
        {
            "code": "low_coverage",
            "short": "LC",
            "severity": severity,
            "message": "called on 14 reads",
        }
    ]


def test_pivots_samples_into_marker_blocks() -> None:
    """The marker is the block and the samples are the rows, not the reverse."""
    payloads = [
        _payload("S1", [_row("S1", "vWA"), _row("S1", "TPOX")]),
        _payload("S2", [_row("S2", "vWA"), _row("S2", "TPOX")]),
    ]

    cohort = build_cohort_payload(payloads)

    assert [b.marker for b in cohort["blocks"]] == ["vWA", "TPOX"], "panel order, not alphabetical"
    assert [r["sample"] for r in cohort["blocks"][0].rows] == ["S1", "S2"]


def test_a_marker_failing_across_the_cohort_is_named() -> None:
    """The signal this view exists for.

    One sample missing a locus is a sample problem. Every sample missing the
    same locus is a panel or assay problem, and a percentage alone would not
    say which locus to look at.
    """
    payloads = [
        _payload(f"S{i}", [_row(f"S{i}", "vWA"), _row(f"S{i}", "DYS391", a1=None, a2=None)])
        for i in range(4)
    ]

    summary = build_cohort_payload(payloads)["summary"]

    named = [m["marker"] for m in summary["worst_markers"]]
    assert named[0] == "DYS391", f"the failing locus must lead the list, got {named}"
    assert "vWA" not in named, "a locus called in every sample is not worth flagging"


def test_flag_counts_are_per_marker() -> None:
    """How many samples raised each code at this locus, commonest first."""
    payloads = [_payload(f"S{i}", [_row(f"S{i}", "TH01", flags=_lc())]) for i in range(3)]

    block = build_cohort_payload(payloads)["blocks"][0]

    assert block.flag_counts == [("LC", "low_coverage", 3)]


def test_sample_names_link_to_their_own_report(tmp_path: Path) -> None:
    """Clicking a sample has to reach the per-locus evidence, which lives there."""
    payloads = [_payload("S1", [_row("S1", "vWA")]), _payload("S2", [_row("S2", "vWA")])]
    out = tmp_path / "cohort.html"

    build_cohort_report(payloads, out, report_hrefs={"S1": "S1/S1.html"})
    doc = lxml.html.fromstring(out.read_text(encoding="utf-8"))

    links = {a.text_content().strip(): a.get("href") for a in doc.cssselect("td.marker a")}
    assert links.get("S1") == "S1/S1.html"
    # S2 had no report written, and a link that 404s is worse than no link.
    assert "S2" not in links


def test_report_carries_alleles_coverage_and_isfg(tmp_path: Path) -> None:
    """The four things asked for, in one row."""
    out = tmp_path / "cohort.html"
    build_cohort_report(
        [_payload("S1", [_row("S1", "vWA", flags=_lc())]), _payload("S2", [_row("S2", "vWA")])],
        out,
    )
    html = out.read_text(encoding="utf-8")

    assert "CE14_TATC[10]" in html, "the ISFG string must be in the table"
    assert ">33<" in html or ">33 <" in html, "the called coverage must be in the table"
    assert ">LC<" in html, "the QC chip must be abbreviated, as in the single-sample view"
    assert 'title="low_coverage: called on 14 reads"' in html, "and expand on hover"


def test_empty_input_is_an_error_not_an_empty_page(tmp_path: Path) -> None:
    """A cohort report with no rows would look like a cohort with no problems."""
    with pytest.raises(FrontstrError, match="nothing to build"):
        build_cohort_payload([_payload("S1", [])])


def test_one_marker_visible_at_a_time(tmp_path: Path) -> None:
    """2,700 rows of continuous scroll is not a table anyone can use.

    Stacking every marker meant 108 samples across 25 loci arrived as one
    unbroken scroll. Only the first block renders visible; the rest carry
    ``is-hidden`` until their tab is chosen.
    """
    out = tmp_path / "cohort.html"
    build_cohort_report(
        [
            _payload("S1", [_row("S1", "vWA"), _row("S1", "TPOX")]),
            _payload("S2", [_row("S2", "vWA"), _row("S2", "TPOX")]),
        ],
        out,
        write_xlsx=False,
    )
    doc = lxml.html.fromstring(out.read_text(encoding="utf-8"))

    tabs = doc.find_class("marker-tab")
    assert [t.get("data-target") for t in tabs] == ["marker-vWA", "marker-TPOX"]
    assert "is-active" in (tabs[0].get("class") or "")

    blocks = doc.find_class("marker-block")
    assert "is-hidden" not in (blocks[0].get("class") or ""), "the first marker must be visible"
    assert "is-hidden" in (blocks[1].get("class") or ""), "the rest must start hidden"


def test_a_tab_badges_a_marker_that_needs_opening(tmp_path: Path) -> None:
    """The strip doubles as the overview, so a badge has to mean something."""
    out = tmp_path / "cohort.html"
    build_cohort_report(
        [
            _payload("S1", [_row("S1", "vWA"), _row("S1", "DYS391", a1=None, a2=None)]),
            _payload("S2", [_row("S2", "vWA"), _row("S2", "DYS391", a1=None, a2=None)]),
        ],
        out,
        write_xlsx=False,
    )
    doc = lxml.html.fromstring(out.read_text(encoding="utf-8"))

    by_marker = {t.get("data-marker"): t for t in doc.find_class("marker-tab")}
    assert "0/2" in by_marker["dys391"].text_content(), "an uncalled locus shows its call rate"
    assert not by_marker["vwa"].find_class("tab-badge"), "a clean locus carries no badge"


def test_workbook_is_written_and_linked(tmp_path: Path) -> None:
    """The download has to work from a report opened off a disk, with no server."""
    import openpyxl

    out = tmp_path / "cohort.html"
    build_cohort_report(
        [_payload("S1", [_row("S1", "vWA", flags=_lc())]), _payload("S2", [_row("S2", "vWA")])],
        out,
    )

    xlsx = tmp_path / "cohort.xlsx"
    assert xlsx.exists(), "no workbook beside the report"
    assert 'href="cohort.xlsx"' in out.read_text(encoding="utf-8")

    wb = openpyxl.load_workbook(xlsx)
    assert wb.sheetnames == ["Genotypes", "By marker", "By sample"]
    genotypes = wb["Genotypes"]
    assert genotypes.max_row == 3, "one header plus one row per (marker, sample)"
    header = [c.value for c in genotypes[1]]
    assert "ISFG allele 1" in header and "Cov" in header
