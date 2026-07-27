"""Tests for the multi-sheet XLSX review workbook.

The workbook is a review instrument, so the tests care about the properties a
reviewer depends on: that nothing is silently dropped, that flagged markers are
visible without hunting, and that the sheet a review starts from is ordered
worst-first.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from frontstr.exports.xlsx import write_run_xlsx


def _allele(
    label: str,
    reads: int,
    *,
    index: int = 0,
    status: str = "allele",
    iso: str | None = None,
    consensus: str = "AGAT" * 10,
) -> dict[str, Any]:
    return {
        "cluster_index": index,
        "consensus": consensus,
        "length_bp": len(consensus),
        "number_label": label,
        "number_method": "period_ce",
        "isfg": f"[AGAT]{label}",
        "status": status,
        "n_reads_total": reads,
        "n_reads_hp1": 0,
        "n_reads_hp2": 0,
        "n_reads_hp_none": reads,
        "n_forward": reads // 2,
        "n_reverse": reads - reads // 2,
        "mean_qual": 30.0,
        "expected_stutter": 0.0,
        "n_reads_absorbed": 0,
        "fraction": 0.5,
        "consensus_method": "poa_spoa",
        "iso": {"suffix": iso, "match_type": "exact" if iso else "none", "is_isoallele": bool(iso)},
        "flags": [],
    }


def _marker(
    name: str,
    alleles: list[dict[str, Any]],
    *,
    called: list[dict[str, Any]] | None = None,
    flags: list[dict[str, str]] | None = None,
) -> dict[str, Any]:
    return {
        "marker_name": name,
        "system": {"motif": "AGAT", "period": 4, "strand": "+", "marker_type": "str"},
        "call_rule": "heterozygous",
        "total_reads": sum(a["n_reads_total"] for a in alleles),
        "alleles": alleles,
        "alleles_called": alleles if called is None else called,
        "flags": flags or [],
    }


def _payload(markers: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "meta": {
            "sample_name": "S1",
            "operator": "op",
            "run_id": "R1",
            "panel_name": "P",
            "panel_version": "1",
            "reference_build": "GRCh38",
        },
        "audit": {
            "tool_version": "0.1",
            "poa_backend": "poa_spoa",
            "stutter_model_version": "m1",
            "stutter_model_protocol": "wgs_pcr_free",
            "analytical_thresh": 0.02,
            "calling_thresh": 0.10,
            "qc_thresholds": {
                "low_coverage_reads": 20,
                "strand_bias_p": 0.01,
                "strand_bias_min_reads": 10,
            },
            "inputs": [{"role": "bam", "path": "s.bam", "sha256": "abc"}],
            "flag_counts": {"low_coverage": 1},
            "flags_checked": ["low_coverage", "dropout"],
            "markers_needing_review": ["M2"],
            "integrity_sha256": "deadbeef",
        },
        "results": markers,
    }


def _sheet_rows(path: Path, name: str) -> list[tuple[Any, ...]]:
    ws = load_workbook(path)[name]
    return list(ws.iter_rows(min_row=2, values_only=True))


def test_workbook_has_the_five_review_sheets(tmp_path: Path) -> None:
    out = write_run_xlsx(_payload([_marker("M1", [_allele("8", 20)])]), tmp_path / "o.xlsx")
    assert load_workbook(out).sheetnames == [
        "Profile",
        "Sequences",
        "Evidence",
        "QC",
        "Audit",
    ]


def test_profile_has_one_row_per_marker(tmp_path: Path) -> None:
    payload = _payload(
        [
            _marker("M1", [_allele("8", 20), _allele("9", 18, index=1)]),
            _marker("M2", [_allele("11", 30)]),
        ]
    )
    out = write_run_xlsx(payload, tmp_path / "o.xlsx")
    rows = _sheet_rows(out, "Profile")
    assert [r[0] for r in rows] == ["M1", "M2"]
    assert rows[0][1] == "8 / 9"
    assert rows[1][1] == "11"


def test_flagged_markers_are_tinted_on_the_profile_sheet(tmp_path: Path) -> None:
    """A reviewer must see which rows need attention without reading every cell."""
    payload = _payload(
        [
            _marker("CLEAN", [_allele("8", 30)]),
            _marker(
                "FLAGGED",
                [_allele("9", 12)],
                flags=[{"code": "low_coverage", "severity": "warn", "message": "thin"}],
            ),
        ]
    )
    out = write_run_xlsx(payload, tmp_path / "o.xlsx")
    ws = load_workbook(out)["Profile"]
    clean_fill = ws.cell(2, 1).fill.fgColor.rgb
    flagged_fill = ws.cell(3, 1).fill.fgColor.rgb
    assert flagged_fill != clean_fill


def test_evidence_keeps_uncalled_clusters(tmp_path: Path) -> None:
    """The suppressed and stutter clusters are the reason a call looks the way it does."""
    called = _allele("8", 20)
    stutter = _allele("7", 3, index=1, status="stutter")
    phantom = _allele("8", 4, index=2, status="hp_phantom")
    payload = _payload([_marker("M1", [called, stutter, phantom], called=[called])])

    out = write_run_xlsx(payload, tmp_path / "o.xlsx")
    rows = _sheet_rows(out, "Evidence")
    assert len(rows) == 3
    statuses = {r[2] for r in rows}
    assert statuses == {"allele", "stutter", "hp_phantom"}
    assert [r[3] for r in rows] == ["yes", None, None]


def test_sequences_sheet_carries_the_full_consensus(tmp_path: Path) -> None:
    consensus = "AGAT" * 25
    payload = _payload([_marker("M1", [_allele("25", 20, consensus=consensus, iso="a")])])
    out = write_run_xlsx(payload, tmp_path / "o.xlsx")
    row = _sheet_rows(out, "Sequences")[0]
    assert row[-1] == consensus, "the sequence must not be truncated into the cell"
    assert row[6] == "a"


def test_qc_sheet_is_ordered_worst_first(tmp_path: Path) -> None:
    payload = _payload(
        [
            _marker(
                "M1",
                [_allele("8", 30)],
                flags=[{"code": "isoallele", "severity": "info", "message": "i"}],
            ),
            _marker(
                "M2",
                [_allele("9", 30)],
                flags=[{"code": "dropout", "severity": "error", "message": "e"}],
            ),
            _marker(
                "M3",
                [_allele("7", 30)],
                flags=[{"code": "low_coverage", "severity": "warn", "message": "w"}],
            ),
        ]
    )
    out = write_run_xlsx(payload, tmp_path / "o.xlsx")
    rows = _sheet_rows(out, "QC")
    assert [r[0] for r in rows] == ["error", "warn", "info"]


def test_qc_sheet_includes_allele_level_flags(tmp_path: Path) -> None:
    allele = _allele("8", 20)
    allele["flags"] = [{"code": "strand_bias", "severity": "warn", "message": "skewed"}]
    out = write_run_xlsx(_payload([_marker("M1", [allele])]), tmp_path / "o.xlsx")
    rows = _sheet_rows(out, "QC")
    assert any("cluster 0" in str(r[1]) for r in rows)


def test_audit_sheet_records_the_run_configuration(tmp_path: Path) -> None:
    out = write_run_xlsx(_payload([_marker("M1", [_allele("8", 20)])]), tmp_path / "o.xlsx")
    pairs = dict(_sheet_rows(out, "Audit"))
    assert pairs["POA backend"] == "poa_spoa"
    assert pairs["Stutter protocol"] == "wgs_pcr_free"
    assert pairs["Low-coverage floor (reads)"] == 20
    assert pairs["Audit record sha256"] == "deadbeef"
    assert pairs["  sha256 (bam)"] == "abc"


def test_audit_sheet_shows_checked_conditions_that_did_not_fire(tmp_path: Path) -> None:
    """A zero is information: it says the condition was evaluated."""
    out = write_run_xlsx(_payload([_marker("M1", [_allele("8", 20)])]), tmp_path / "o.xlsx")
    pairs = dict(_sheet_rows(out, "Audit"))
    assert pairs["Flag: low_coverage"] == 1
    assert pairs["Flag: dropout"] == 0


def test_sheets_are_frozen_and_filterable(tmp_path: Path) -> None:
    out = write_run_xlsx(_payload([_marker("M1", [_allele("8", 20)])]), tmp_path / "o.xlsx")
    wb = load_workbook(out)
    for name in ("Profile", "Sequences", "Evidence", "QC"):
        assert wb[name].freeze_panes == "A2"
        assert wb[name].auto_filter.ref, f"{name} should be filterable"


def test_empty_run_still_produces_a_readable_workbook(tmp_path: Path) -> None:
    out = write_run_xlsx(_payload([]), tmp_path / "o.xlsx")
    wb = load_workbook(out)
    assert wb["Profile"].max_row == 1, "header only"
    assert wb["Audit"].max_row > 1, "configuration is still worth recording"
